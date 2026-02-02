# License: https://github.com/LauGroup/EncDE/blob/main/LICENSE
# Author: Rezwan Siddiquee; rezwan.siddiquee@sydney.edu.au
# School of Chemistry, The University of Sydney

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from pathlib import Path

# --- User Defined Parameters ---
# Directory containing the Excel files ('.' for current dir)
INPUT_DIRECTORY = Path('Demo/')
# Folder to save the modified Excel files (can be the same as INPUT_DIRECTORY if overwriting)
OUTPUT_DIRECTORY = Path('Demo/')
# --- End User Defined Parameters ---

# Function to count deletions in a sequence
def count_deletions(sequence):
    return sequence.count("-")

# Function to apply Courier font to specific columns (e.g., B and C)
def apply_courier_font_to_columns(sheet, columns):
    courier_font = Font(name='Courier New')
    for col in columns:
        for cell in sheet[col]:
            cell.font = courier_font

# Create output directory if it doesn't exist
if not OUTPUT_DIRECTORY.exists():
    OUTPUT_DIRECTORY.mkdir(parents=True)

# Loop through all Excel files in the INPUT_DIRECTORY
for filename in os.listdir(INPUT_DIRECTORY):
    if filename.endswith(".xlsx"):
        input_file_path = INPUT_DIRECTORY / filename
        # If output is different, define output path, otherwise overwrite
        if INPUT_DIRECTORY == OUTPUT_DIRECTORY:
            output_file_path = input_file_path
        else:
            output_file_path = OUTPUT_DIRECTORY / filename

        try:
            df = pd.read_excel(input_file_path)

            # Ensure the WT sequence is in the first row (column 2, zero-indexed)
            if len(df.columns) > 2 and not df.empty:
                # Assuming the third column (index 2) contains the sequences
                wt_sequence = str(df.iloc[0, 2])  # Get the WT sequence from the first row of column 2
                wt_deletions = count_deletions(wt_sequence)  # Count deletions in WT sequence

                # Function to calculate the number of deletions compared to WT
                def calculate_deletions(sequence):
                    seq_deletions = count_deletions(str(sequence)) # Ensure sequence is string
                    return seq_deletions - wt_deletions

                # Add the new column as the last column in the DataFrame
                # Apply to the third column (index 2)
                df['# of Deletions'] = df.iloc[:, 2].apply(calculate_deletions)

                # Save the modified DataFrame
                df.to_excel(output_file_path, index=False)

            # Re-open the workbook to apply font changes (if it was saved)
            if os.path.exists(output_file_path):
                wb = load_workbook(output_file_path)
                if 'Sheet1' in wb.sheetnames:
                    sheet = wb['Sheet1']  # Assumes the sheet name is 'Sheet1', adjust if necessary
                    # Apply Courier font to columns B and C
                    apply_courier_font_to_columns(sheet, ['B', 'C'])
                    wb.save(output_file_path) # Save the workbook with font changes
                else:
                    print(f"Warning: 'Sheet1' not found in {output_file_path}. Font not applied.")
            else:
                print(f"Warning: File {output_file_path} not found for font changes. It might not have been processed or saved correctly.")


        except Exception as e:
            print(f"Error processing file {filename}: {e}")


print(f"Excel files in '{INPUT_DIRECTORY}' processed. Modified files saved to '{OUTPUT_DIRECTORY}'.")
print("Deletion counts added and Courier font applied to columns B and C where applicable.")